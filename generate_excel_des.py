"""
generate_excel_des.py
=====================
Script untuk menghasilkan file Excel perhitungan manual
Double Exponential Smoothing (DES) / Holt's Linear Trend
dari data historis database MySQL sistem ALS Energy.

Cara pakai:
  1. Jalankan dari folder monitoring-listrik:
     python generate_excel_des.py
  2. File 'Perhitungan_DES_ALS_Energy.xlsx' akan dibuat di folder yang sama.
"""

import os
import sys
import itertools
import json
import datetime

# ─── Cek dependensi ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("❌  openpyxl belum terinstall.")
    print("    Jalankan:  pip install openpyxl")
    sys.exit(1)

try:
    import mysql.connector
except ImportError:
    print("❌  mysql-connector-python belum terinstall.")
    print("    Jalankan:  pip install mysql-connector-python")
    sys.exit(1)

# ─── Konfigurasi ─────────────────────────────────────────────────────────────────
CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config', 'settings.json')

# Rentang tanggal dataset pengujian (sesuai laporan skripsi)
DATE_START = "2026-04-13"
DATE_END   = "2026-05-22"

# Mesin yang diuji  ('all' / 'mesin1' / 'mesin2')
MACHINE_ID = "all"

# Parameter DES optimal (akan di-override oleh hasil trial & error jika AUTOFIT=True)
AUTOFIT    = True   # True = sistem cari sendiri α β terbaik (81 kombinasi)
ALPHA_FIX  = 0.4   # dipakai jika AUTOFIT=False
BETA_FIX   = 0.2   # dipakai jika AUTOFIT=False

# Periode proyeksi ke depan
FORECAST_PERIOD = 7   # hari

OUTPUT_FILE = "Perhitungan_DES_ALS_Energy.xlsx"

# ─── Palet Warna ─────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # Biru tua (header utama)
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "2E6DAD"   # Biru sedang (sub-header)
C_SUBHEAD_FG  = "FFFFFF"
C_ALT1        = "EBF3FB"   # Biru sangat muda (baris ganjil)
C_ALT2        = "FFFFFF"   # Putih (baris genap)
C_ACTUAL      = "1A5276"   # Warna teks data aktual
C_FORECAST    = "1D6A39"   # Warna teks forecast
C_GOOD        = "D5F5E3"   # Hijau muda (MAPE bagus)
C_WARN        = "FEF9E7"   # Kuning muda (MAPE cukup)
C_BAD         = "FADBD8"   # Merah muda (MAPE buruk)
C_SECTION     = "D6EAF8"   # Biru muda (judul section)
C_FUTURE      = "FFF3CD"   # Kuning muda (baris proyeksi)
C_FUTURE_FG   = "7D6608"
C_INIT        = "F2F3F4"   # Abu-abu (baris inisialisasi)

# ─── Fungsi Helper ───────────────────────────────────────────────────────────────

def get_db():
    with open(CONFIG_PATH) as f:
        cfg = json.load(f)
    return mysql.connector.connect(
        host=cfg.get('db_host', 'localhost'),
        user=cfg.get('db_user', 'root'),
        password=cfg.get('db_password', ''),
        database=cfg.get('db_name', 'als_energy')
    )

def fetch_data(date_start, date_end, machine_id):
    """Ambil data historis kWh harian dari database."""
    conn = get_db()
    cur  = conn.cursor(dictionary=True)

    conds  = ["DATE(timestamp) >= %s", "DATE(timestamp) <= %s"]
    params = [date_start, date_end]
    if machine_id != 'all':
        conds.append("mesin_id = %s")
        params.append(machine_id)

    where = "WHERE " + " AND ".join(conds)
    query = f"""
        SELECT DATE(timestamp) AS dt, mesin_id, AVG(daya) AS avg_w
        FROM sensor_data
        {where}
        GROUP BY DATE(timestamp), mesin_id
        ORDER BY dt ASC
    """
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Jumlahkan rata-rata daya tiap mesin per tanggal → kWh/hari
    kwh_by_date = {}
    for row in rows:
        ds = row['dt'].strftime('%Y-%m-%d')
        kwh_by_date[ds] = kwh_by_date.get(ds, 0) + (row['avg_w'] * 24 / 1000.0)

    dates   = sorted(kwh_by_date.keys())
    actuals = [round(kwh_by_date[d], 4) for d in dates]
    return dates, actuals

def run_des(actuals, alpha, beta):
    """Jalankan DES dan kembalikan (s, b, forecast, mape)."""
    n = len(actuals)
    s = [0.0] * n
    b = [0.0] * n
    f = [0.0] * n

    s[0] = actuals[0]
    b[0] = actuals[1] - actuals[0]
    f[0] = actuals[0]
    if n > 1:
        f[1] = s[0] + b[0]

    for t in range(1, n):
        s[t] = alpha * actuals[t] + (1 - alpha) * (s[t-1] + b[t-1])
        b[t] = beta  * (s[t] - s[t-1]) + (1 - beta) * b[t-1]
        if t + 1 < n:
            f[t+1] = s[t] + b[t]

    apes = [abs((actuals[t] - f[t]) / actuals[t]) * 100
            for t in range(2, n) if actuals[t] != 0]
    mape = sum(apes) / len(apes) if apes else 0.0
    return s, b, f, mape

def find_best(actuals):
    """Cari α β terbaik dari 81 kombinasi."""
    vals = [round(x * 0.1, 1) for x in range(1, 10)]
    best = {'mape': float('inf'), 'alpha': 0.4, 'beta': 0.2}
    results = []
    for alpha, beta in itertools.product(vals, vals):
        _, _, _, mape = run_des(actuals, alpha, beta)
        results.append((alpha, beta, mape))
        if mape < best['mape']:
            best = {'mape': mape, 'alpha': alpha, 'beta': beta}
    return best, results

def thin_border():
    side = Side(style='thin', color='AAAAAA')
    return Border(left=side, right=side, top=side, bottom=side)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(bold=True, color=C_HEADER_FG, size=11):
    return Font(name='Calibri', bold=bold, color=color, size=size)

def write_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if num_fmt: cell.number_format = num_fmt
    return cell

# ─── MAIN ────────────────────────────────────────────────────────────────────────

def main():
    print("[INFO] Menghubungkan ke database MySQL ...")
    try:
        dates, actuals = fetch_data(DATE_START, DATE_END, MACHINE_ID)
    except Exception as e:
        print(f"❌  Gagal koneksi database: {e}")
        print("    Pastikan server MySQL berjalan dan settings.json sudah benar.")
        sys.exit(1)

    n = len(actuals)
    if n < 3:
        print(f"❌  Data terlalu sedikit ({n} hari). Minimal 3 hari diperlukan.")
        sys.exit(1)

    print(f"[OK]   Data ditemukan: {n} hari ({dates[0]} s/d {dates[-1]})")

    # Cari parameter optimal
    if AUTOFIT:
        print("[INFO] Menjalankan trial & error 81 kombinasi alpha x beta ...")
        best, all_combo = find_best(actuals)
        alpha = best['alpha']
        beta  = best['beta']
        print(f"[OK]   Parameter optimal: alpha={alpha}, beta={beta}, MAPE={best['mape']:.4f}%")
    else:
        alpha = ALPHA_FIX
        beta  = BETA_FIX
        _, _, _, mape = run_des(actuals, alpha, beta)
        best = {'alpha': alpha, 'beta': beta, 'mape': mape}
        all_combo = [(alpha, beta, mape)]

    # Jalankan DES dengan parameter terbaik
    s_vals, b_vals, f_vals, final_mape = run_des(actuals, alpha, beta)

    # Proyeksi ke depan
    future_dates, future_vals = [], []
    last_date = datetime.datetime.strptime(dates[-1], '%Y-%m-%d')
    for m in range(1, FORECAST_PERIOD + 1):
        fv = max(s_vals[-1] + m * b_vals[-1], 0.01)
        fd = (last_date + datetime.timedelta(days=m)).strftime('%Y-%m-%d')
        future_dates.append(fd)
        future_vals.append(round(fv, 4))

    # ─── Buat Workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Tabel Perhitungan DES
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Perhitungan DES"

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center')
    RIGHT  = Alignment(horizontal='right',  vertical='center')

    # ─── Judul Dokumen ────────────────────────────────────────────────────────
    ws1.merge_cells('A1:L1')
    c = ws1['A1']
    c.value     = "PERHITUNGAN MANUAL DOUBLE EXPONENTIAL SMOOTHING (DES) / HOLT'S LINEAR TREND"
    c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=14)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:L2')
    c = ws1['A2']
    c.value     = "Sistem Monitoring Energi Listrik Berbasis IoT – ALS Energy"
    c.font      = Font(name='Calibri', italic=True, color=C_HEADER_FG, size=11)
    c.fill      = make_fill(C_SUBHEAD_BG)
    c.alignment = CENTER
    ws1.row_dimensions[2].height = 18

    # ─── Info Dataset ─────────────────────────────────────────────────────────
    info = [
        ("Dataset",          f"{dates[0]} s/d {dates[-1]}  ({n} hari)"),
        ("Mesin",            "Semua Mesin (Mesin 1 + Mesin 2)" if MACHINE_ID == 'all' else MACHINE_ID.capitalize()),
        ("Metode",           "Double Exponential Smoothing (Holt's Linear Trend)"),
        ("Kombinasi Diuji",  "81 kombinasi (α: 0.1–0.9, β: 0.1–0.9)"),
        ("Parameter Optimal",f"α (Alpha) = {alpha}  |  β (Beta) = {beta}"),
        ("MAPE Terbaik",     f"{final_mape:.4f}%  ({'Baik ✔' if final_mape < 20 else 'Cukup' if final_mape < 50 else 'Buruk'})"),
        ("Proyeksi",         f"{FORECAST_PERIOD} hari ke depan"),
    ]

    r = 4
    ws1.merge_cells(f'A{r}:D{r}')
    c = ws1.cell(r, 1, "INFORMASI DATASET & PARAMETER")
    c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=11)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = CENTER
    ws1.merge_cells(f'E{r}:L{r}')
    ws1.cell(r, 5).fill = make_fill(C_HEADER_BG)

    for label, val in info:
        r += 1
        ws1.merge_cells(f'A{r}:D{r}')
        c1 = ws1.cell(r, 1, label)
        c1.font      = Font(name='Calibri', bold=True, color='1A1A1A', size=10)
        c1.fill      = make_fill(C_ALT1)
        c1.alignment = LEFT
        c1.border    = thin_border()

        ws1.merge_cells(f'E{r}:L{r}')
        c2 = ws1.cell(r, 5, val)
        c2.font      = Font(name='Calibri', color='1A1A1A', size=10)
        c2.fill      = make_fill(C_ALT2)
        c2.alignment = LEFT
        c2.border    = thin_border()

    # ─── Header Tabel Utama ───────────────────────────────────────────────────
    r += 2
    HEADER_ROW = r
    headers = [
        ("No.",           5),
        ("Tanggal",       14),
        ("Data Aktual\n(kWh)", 14),
        ("Level\nSₜ",    14),
        ("Tren\nbₜ",     14),
        ("Forecast\nFₜ", 14),
        ("Error\n(Xₜ – Fₜ)", 14),
        ("Abs. Error\n|Error|", 14),
        ("APE\n(%)",     12),
        ("Keterangan",   20),
    ]

    col_start = 1
    col_map   = {}   # nama → indeks kolom
    for i, (hdr, w) in enumerate(headers):
        col = col_start + i
        ws1.column_dimensions[get_column_letter(col)].width = w
        c = ws1.cell(r, col, hdr)
        c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=10)
        c.fill      = make_fill(C_HEADER_BG)
        c.alignment = CENTER
        c.border    = thin_border()
        col_map[hdr.split('\n')[0]] = col

    ws1.row_dimensions[r].height = 32

    # ─── Isi Baris Data ───────────────────────────────────────────────────────
    def keterangan(t):
        if t == 0: return "Inisialisasi  S₁ = X₁"
        if t == 1: return "Inisialisasi  b₁ = X₂ – X₁"
        return f"Sₜ=α·Xₜ+(1-α)(Sₜ₋₁+bₜ₋₁) ; bₜ=β(Sₜ-Sₜ₋₁)+(1-β)bₜ₋₁"

    ape_values = []   # untuk MAPE rekapitulasi
    for t in range(n):
        r += 1
        bg = C_INIT if t < 2 else (C_ALT1 if t % 2 == 0 else C_ALT2)

        actual = actuals[t]
        s_v    = round(s_vals[t], 4)
        b_v    = round(b_vals[t], 4)
        f_v    = round(f_vals[t], 4) if t >= 2 else "-"
        err    = round(actual - f_vals[t], 4) if t >= 2 else "-"
        abs_e  = round(abs(actual - f_vals[t]), 4) if t >= 2 else "-"
        ape    = round(abs((actual - f_vals[t]) / actual) * 100, 4) if (t >= 2 and actual != 0) else "-"

        if isinstance(ape, float):
            ape_values.append(ape)

        row_data = [t + 1, dates[t], actual, s_v, b_v, f_v, err, abs_e, ape, keterangan(t)]
        for col_i, val in enumerate(row_data):
            col = col_start + col_i
            is_num = isinstance(val, float) and val != "-"
            cell = ws1.cell(r, col, val)
            cell.fill      = make_fill(bg)
            cell.border    = thin_border()
            cell.alignment = CENTER if col_i != 9 else LEFT
            if col_i == 2:  # Aktual
                cell.font = Font(name='Calibri', bold=True, color=C_ACTUAL, size=10)
                if is_num: cell.number_format = '0.0000'
            elif col_i == 5:  # Forecast
                cell.font = Font(name='Calibri', bold=True, color=C_FORECAST, size=10)
                if is_num: cell.number_format = '0.0000'
            elif col_i in (3, 4, 6, 7):
                cell.font = Font(name='Calibri', size=10)
                if is_num: cell.number_format = '0.0000'
            elif col_i == 8:  # APE
                cell.font = Font(name='Calibri', size=10)
                if is_num: cell.number_format = '0.0000'
            else:
                cell.font = Font(name='Calibri', size=10)

        ws1.row_dimensions[r].height = 16

    # ─── Baris MAPE ───────────────────────────────────────────────────────────
    r += 1
    ws1.merge_cells(f'A{r}:H{r}')
    c = ws1.cell(r, 1, "MAPE (Mean Absolute Percentage Error)")
    mape_color = C_GOOD if final_mape < 20 else C_WARN if final_mape < 50 else C_BAD
    c.font      = Font(name='Calibri', bold=True, size=11, color='1A1A1A')
    c.fill      = make_fill(mape_color[0:6] if len(mape_color)==7 else mape_color)
    c.alignment = CENTER
    c.border    = thin_border()

    c2 = ws1.cell(r, 9, round(final_mape, 4))
    c2.font          = Font(name='Calibri', bold=True, size=11)
    c2.number_format = '0.0000"%"'
    c2.fill          = make_fill(mape_color)
    c2.alignment     = CENTER
    c2.border        = thin_border()

    ws1.merge_cells(f'J{r}:L{r}')   # kolom J (index 10)
    c3 = ws1.cell(r, 10, f"{'Baik (< 20%)' if final_mape < 20 else 'Cukup (20-50%)' if final_mape < 50 else 'Buruk (> 50%)'}")
    c3.font      = Font(name='Calibri', bold=True, size=11)
    c3.fill      = make_fill(mape_color)
    c3.alignment = CENTER
    c3.border    = thin_border()
    ws1.row_dimensions[r].height = 20

    # ─── Baris Proyeksi Masa Depan ────────────────────────────────────────────
    r += 2
    ws1.merge_cells(f'A{r}:J{r}')
    c = ws1.cell(r, 1, f"PROYEKSI {FORECAST_PERIOD} HARI KE DEPAN")
    c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=11)
    c.fill      = make_fill("7D6608")
    c.alignment = CENTER
    ws1.row_dimensions[r].height = 20

    for m, (fd, fv) in enumerate(zip(future_dates, future_vals)):
        r += 1
        bg = C_FUTURE
        row_data = [n + m + 1, fd, "-", "-", "-", fv, "-", "-", "-", "Proyeksi ke depan"]
        for col_i, val in enumerate(row_data):
            col  = col_start + col_i
            cell = ws1.cell(r, col, val)
            cell.fill      = make_fill(bg)
            cell.border    = thin_border()
            cell.alignment = CENTER if col_i != 9 else LEFT
            if col_i == 5:
                cell.font          = Font(name='Calibri', bold=True, color=C_FUTURE_FG, size=10)
                cell.number_format = '0.0000'
            else:
                cell.font = Font(name='Calibri', color=C_FUTURE_FG, size=10)
        ws1.row_dimensions[r].height = 16

    # Freeze pane di bawah header tabel
    ws1.freeze_panes = f'A{HEADER_ROW + 1}'
    ws1.sheet_view.showGridLines = True

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Tabel 81 Kombinasi Alpha Beta
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("81 Kombinasi α×β")

    ws2.merge_cells('A1:D1')
    c = ws2['A1']
    c.value     = "HASIL UJI COBA 81 KOMBINASI PARAMETER α × β"
    c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=13)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = CENTER
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells('A2:D2')
    c2 = ws2['A2']
    c2.value     = f"Parameter Optimal Terpilih: α = {alpha}  |  β = {beta}  |  MAPE = {final_mape:.4f}%"
    c2.font      = Font(name='Calibri', bold=True, color=C_SUBHEAD_FG, size=11)
    c2.fill      = make_fill(C_SUBHEAD_BG)
    c2.alignment = CENTER
    ws2.row_dimensions[2].height = 18

    hdrs2 = ["No.", "Alpha (α)", "Beta (β)", "MAPE (%)"]
    widths2 = [6, 14, 14, 16]
    for i, (h, w) in enumerate(zip(hdrs2, widths2)):
        col = i + 1
        ws2.column_dimensions[get_column_letter(col)].width = w
        c = ws2.cell(4, col, h)
        c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=10)
        c.fill      = make_fill(C_HEADER_BG)
        c.alignment = CENTER
        c.border    = thin_border()

    # Urutkan berdasarkan MAPE
    all_combo_sorted = sorted(all_combo, key=lambda x: x[2])
    for idx, (a, b, mape_val) in enumerate(all_combo_sorted):
        row = idx + 5
        is_best = (a == alpha and b == beta)
        bg = C_GOOD if is_best else (C_ALT1 if idx % 2 == 0 else C_ALT2)
        vals2 = [idx + 1, a, b, round(mape_val, 4)]
        for col_i, v in enumerate(vals2):
            cell = ws2.cell(row, col_i + 1, v)
            cell.fill      = make_fill(bg)
            cell.border    = thin_border()
            cell.alignment = CENTER
            if col_i == 3:
                cell.number_format = '0.0000'
                cell.font = Font(name='Calibri', bold=is_best, size=10,
                                 color='1D6A39' if is_best else '1A1A1A')
            else:
                cell.font = Font(name='Calibri', bold=is_best, size=10)
        ws2.row_dimensions[row].height = 15

    ws2.freeze_panes = 'A5'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Rumus & Keterangan
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Rumus & Keterangan")

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 60

    ws3.merge_cells('A1:B1')
    c = ws3['A1']
    c.value     = "RUMUS DOUBLE EXPONENTIAL SMOOTHING (HOLT'S LINEAR TREND)"
    c.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=13)
    c.fill      = make_fill(C_HEADER_BG)
    c.alignment = CENTER
    ws3.row_dimensions[1].height = 24

    rumus = [
        ("", ""),
        ("INISIALISASI", ""),
        ("Level Awal  S₁",  "S₁ = X₁   (data aktual hari pertama)"),
        ("Tren Awal   b₁",  "b₁ = X₂ – X₁"),
        ("", ""),
        ("REKURSI (t = 2, 3, …, n)", ""),
        ("Level  Sₜ",    "Sₜ  = α · Xₜ  +  (1 – α) · (Sₜ₋₁ + bₜ₋₁)"),
        ("Tren   bₜ",    "bₜ  = β · (Sₜ – Sₜ₋₁)  +  (1 – β) · bₜ₋₁"),
        ("Forecast  Fₜ₊₁", "Fₜ₊₁ = Sₜ + bₜ"),
        ("", ""),
        ("PROYEKSI m LANGKAH", ""),
        ("Forecast  Fₜ₊ₘ", "Fₜ₊ₘ = Sₜ + bₜ · m"),
        ("", ""),
        ("EVALUASI", ""),
        ("Error", "Errorₜ = Xₜ – Fₜ"),
        ("Absolute Error", "AEₜ = |Xₜ – Fₜ|"),
        ("APE", "APEₜ = |Xₜ – Fₜ| / Xₜ × 100%"),
        ("MAPE", "MAPE = (1/n) · Σ APEₜ"),
        ("", ""),
        ("KATEGORI MAPE", ""),
        ("Sangat Baik",  "MAPE < 10%"),
        ("Baik",         "10% ≤ MAPE < 20%"),
        ("Cukup",        "20% ≤ MAPE < 50%"),
        ("Buruk",        "MAPE ≥ 50%"),
        ("", ""),
        ("KETERANGAN VARIABEL", ""),
        ("α (Alpha)",   "Koefisien pemulusan level  (0 < α < 1)"),
        ("β (Beta)",    "Koefisien pemulusan tren   (0 < β < 1)"),
        ("Xₜ",          "Data aktual pada waktu ke-t (kWh)"),
        ("Sₜ",          "Nilai level (smoothed level) pada waktu ke-t"),
        ("bₜ",          "Nilai tren (smoothed trend) pada waktu ke-t"),
        ("Fₜ",          "Nilai forecast/peramalan pada waktu ke-t"),
        ("n",           "Jumlah data historis (data latih)"),
        ("m",           "Jangkauan proyeksi ke depan (hari)"),
    ]

    section_labels = {"INISIALISASI", "REKURSI (t = 2, 3, …, n)",
                      "PROYEKSI m LANGKAH", "EVALUASI",
                      "KATEGORI MAPE", "KETERANGAN VARIABEL"}

    for ri, (label, formula) in enumerate(rumus):
        row = ri + 2
        ws3.row_dimensions[row].height = 16
        is_section = label in section_labels
        is_empty   = label == "" and formula == ""

        c1 = ws3.cell(row, 1, label)
        c2 = ws3.cell(row, 2, formula)

        if is_section:
            ws3.merge_cells(f'A{row}:B{row}')
            c1.font      = Font(name='Calibri', bold=True, color=C_HEADER_FG, size=11)
            c1.fill      = make_fill(C_SUBHEAD_BG)
            c1.alignment = LEFT
        elif not is_empty:
            c1.font      = Font(name='Calibri', bold=True, size=10)
            c1.fill      = make_fill(C_ALT1)
            c1.alignment = LEFT
            c2.font      = Font(name='Calibri', size=10)
            c2.fill      = make_fill(C_ALT2)
            c2.alignment = LEFT
            c1.border    = thin_border()
            c2.border    = thin_border()

    # ─── Simpan File ─────────────────────────────────────────────────────────
    output_path = os.path.join(os.path.dirname(__file__), OUTPUT_FILE)
    wb.save(output_path)
    print(f"\n[DONE] File Excel berhasil dibuat!")
    print(f"       Lokasi: {output_path}")
    print(f"\n       Sheet yang tersedia:")
    print(f"       1. Perhitungan DES   - Tabel iterasi lengkap + proyeksi")
    print(f"       2. 81 Kombinasi AxB  - Semua hasil trial & error (diurutkan MAPE)")
    print(f"       3. Rumus & Keterangan - Penjelasan formula DES")
    print(f"\n       Parameter Optimal: alpha={alpha}, beta={beta}, MAPE={final_mape:.4f}%")

if __name__ == '__main__':
    main()
